import io

from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.analytics.schemas import (
    AnalyticsCharts,
    AnalyticsDashboardResponse,
    AttendanceByActivity,
    EnrollmentByDiscipline,
    EnrollmentOverTime,
)
from app.auth.models import User, UserRole
from app.catalog.models import Activity, Enrollment, EnrollmentStatus


async def get_dashboard(
    db: AsyncSession,
    discipline_id: str | None = None,
) -> AnalyticsDashboardResponse:
    # --- total_enrollments ---
    enroll_q = (
        select(func.count())
        .select_from(Enrollment)
        .where(Enrollment.status == EnrollmentStatus.active.value)
    )
    if discipline_id:
        enroll_q = (
            select(func.count())
            .select_from(Enrollment)
            .join(Activity, Enrollment.activity_id == Activity.id)
            .where(
                Enrollment.status == EnrollmentStatus.active.value,
                Activity.discipline_id == discipline_id,
            )
        )
    total_enrollments: int = (await db.scalar(enroll_q)) or 0

    # --- avg_attendance_rate ---
    # AVG of (present/total_per_session * 100) across all sessions
    # Raw SQL via text for the avg attendance calculation
    if discipline_id:
        avg_query = text("""
            SELECT
                AVG(sub.rate)
            FROM (
                SELECT
                    cs.id,
                    CASE WHEN COUNT(ar.id) = 0 THEN 0
                         ELSE SUM(CASE WHEN ar.present THEN 1 ELSE 0 END)::float
                              / COUNT(ar.id) * 100
                    END AS rate
                FROM class_sessions cs
                JOIN activities a ON cs.activity_id = a.id
                LEFT JOIN attendance_records ar ON ar.class_session_id = cs.id
                WHERE a.discipline_id = :discipline_id
                GROUP BY cs.id
            ) sub
        """)
        avg_result = await db.execute(avg_query, {"discipline_id": discipline_id})
    else:
        avg_query = text("""
            SELECT
                AVG(sub.rate)
            FROM (
                SELECT
                    cs.id,
                    CASE WHEN COUNT(ar.id) = 0 THEN 0
                         ELSE SUM(CASE WHEN ar.present THEN 1 ELSE 0 END)::float
                              / COUNT(ar.id) * 100
                    END AS rate
                FROM class_sessions cs
                LEFT JOIN attendance_records ar ON ar.class_session_id = cs.id
                GROUP BY cs.id
            ) sub
        """)
        avg_result = await db.execute(avg_query)

    avg_val = avg_result.scalar()
    avg_attendance_rate: float = (
        round(float(avg_val), 2) if avg_val is not None else 0.0
    )

    # --- active_activities ---
    if discipline_id:
        act_q = (
            select(func.count())
            .select_from(Activity)
            .where(Activity.discipline_id == discipline_id)
        )
    else:
        act_q = select(func.count()).select_from(Activity)
    active_activities: int = (await db.scalar(act_q)) or 0

    # --- total_students ---
    students_q = select(func.count()).select_from(User).where(User.role == UserRole.alumno)
    total_students: int = (await db.scalar(students_q)) or 0

    # --- enrollments_over_time (last 8 weeks, grouped by week) ---
    if discipline_id:
        ot_query = text("""
            SELECT
                TO_CHAR(DATE_TRUNC('week', e.enrolled_at), 'YYYY-MM-DD') AS week,
                COUNT(*) AS cnt
            FROM enrollments e
            JOIN activities a ON e.activity_id = a.id
            WHERE e.status = 'active'
              AND a.discipline_id = :discipline_id
              AND e.enrolled_at >= NOW() - INTERVAL '8 weeks'
            GROUP BY DATE_TRUNC('week', e.enrolled_at)
            ORDER BY DATE_TRUNC('week', e.enrolled_at)
        """)
        ot_result = await db.execute(ot_query, {"discipline_id": discipline_id})
    else:
        ot_query = text("""
            SELECT
                TO_CHAR(DATE_TRUNC('week', e.enrolled_at), 'YYYY-MM-DD') AS week,
                COUNT(*) AS cnt
            FROM enrollments e
            WHERE e.status = 'active'
              AND e.enrolled_at >= NOW() - INTERVAL '8 weeks'
            GROUP BY DATE_TRUNC('week', e.enrolled_at)
            ORDER BY DATE_TRUNC('week', e.enrolled_at)
        """)
        ot_result = await db.execute(ot_query)

    enrollments_over_time = [
        EnrollmentOverTime(date=str(row[0]), count=int(row[1]))
        for row in ot_result.all()
    ]

    # --- attendance_by_activity ---
    if discipline_id:
        aba_query = text("""
            SELECT
                a.name AS activity_name,
                AVG(
                    CASE WHEN total_per_session.total = 0 THEN 0
                         ELSE present_per_session.present_count::float
                              / total_per_session.total * 100
                    END
                ) AS avg_rate,
                COUNT(DISTINCT cs.id) AS total_sessions
            FROM activities a
            JOIN class_sessions cs ON cs.activity_id = a.id
            JOIN (
                SELECT class_session_id, COUNT(*) AS total
                FROM attendance_records
                GROUP BY class_session_id
            ) total_per_session ON total_per_session.class_session_id = cs.id
            JOIN (
                SELECT class_session_id,
                       SUM(CASE WHEN present THEN 1 ELSE 0 END) AS present_count
                FROM attendance_records
                GROUP BY class_session_id
            ) present_per_session ON present_per_session.class_session_id = cs.id
            WHERE a.discipline_id = :discipline_id
            GROUP BY a.id, a.name
            HAVING COUNT(DISTINCT cs.id) >= 1
            ORDER BY a.name
        """)
        aba_result = await db.execute(aba_query, {"discipline_id": discipline_id})
    else:
        aba_query = text("""
            SELECT
                a.name AS activity_name,
                AVG(
                    CASE WHEN total_per_session.total = 0 THEN 0
                         ELSE present_per_session.present_count::float
                              / total_per_session.total * 100
                    END
                ) AS avg_rate,
                COUNT(DISTINCT cs.id) AS total_sessions
            FROM activities a
            JOIN class_sessions cs ON cs.activity_id = a.id
            JOIN (
                SELECT class_session_id, COUNT(*) AS total
                FROM attendance_records
                GROUP BY class_session_id
            ) total_per_session ON total_per_session.class_session_id = cs.id
            JOIN (
                SELECT class_session_id,
                       SUM(CASE WHEN present THEN 1 ELSE 0 END) AS present_count
                FROM attendance_records
                GROUP BY class_session_id
            ) present_per_session ON present_per_session.class_session_id = cs.id
            GROUP BY a.id, a.name
            HAVING COUNT(DISTINCT cs.id) >= 1
            ORDER BY a.name
        """)
        aba_result = await db.execute(aba_query)

    attendance_by_activity = [
        AttendanceByActivity(
            activity_name=str(row[0]),
            avg_rate=round(float(row[1]), 2) if row[1] is not None else 0.0,
            total_sessions=int(row[2]),
        )
        for row in aba_result.all()
    ]

    # --- enrollments_by_discipline ---
    if discipline_id:
        ebd_query = text("""
            SELECT d.name, COUNT(e.id)
            FROM disciplines d
            JOIN activities a ON a.discipline_id = d.id
            JOIN enrollments e ON e.activity_id = a.id
            WHERE e.status = 'active'
              AND d.id = :discipline_id
            GROUP BY d.id, d.name
            ORDER BY d.name
        """)
        ebd_result = await db.execute(ebd_query, {"discipline_id": discipline_id})
    else:
        ebd_query = text("""
            SELECT d.name, COUNT(e.id)
            FROM disciplines d
            JOIN activities a ON a.discipline_id = d.id
            JOIN enrollments e ON e.activity_id = a.id
            WHERE e.status = 'active'
            GROUP BY d.id, d.name
            ORDER BY d.name
        """)
        ebd_result = await db.execute(ebd_query)

    enrollments_by_discipline = [
        EnrollmentByDiscipline(discipline_name=str(row[0]), count=int(row[1]))
        for row in ebd_result.all()
    ]

    charts = AnalyticsCharts(
        enrollments_over_time=enrollments_over_time,
        attendance_by_activity=attendance_by_activity,
        enrollments_by_discipline=enrollments_by_discipline,
    )

    return AnalyticsDashboardResponse(
        total_enrollments=total_enrollments,
        avg_attendance_rate=avg_attendance_rate,
        active_activities=active_activities,
        total_students=total_students,
        charts=charts,
    )


async def export_dashboard(
    db: AsyncSession,
    discipline_id: str | None = None,
) -> bytes:
    import openpyxl

    dashboard = await get_dashboard(db, discipline_id=discipline_id)

    wb = openpyxl.Workbook()

    # Hoja 1: Inscripciones (over time)
    ws1 = wb.active
    ws1.title = "Inscripciones"
    ws1.append(["Semana", "Cantidad"])
    for item in dashboard.charts.enrollments_over_time:
        ws1.append([item.date, item.count])

    # Hoja 2: Asistencia (by activity)
    ws2 = wb.create_sheet("Asistencia")
    ws2.append(["Actividad", "Tasa Promedio (%)", "Total Sesiones"])
    for item in dashboard.charts.attendance_by_activity:
        ws2.append([item.activity_name, item.avg_rate, item.total_sessions])

    # Hoja 3: Por Disciplina
    ws3 = wb.create_sheet("Por Disciplina")
    ws3.append(["Disciplina", "Inscripciones"])
    for item in dashboard.charts.enrollments_by_discipline:
        ws3.append([item.discipline_name, item.count])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
